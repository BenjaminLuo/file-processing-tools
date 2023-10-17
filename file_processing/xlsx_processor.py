from file_processor_strategy import FileProcessorStrategy
from openpyxl import load_workbook 
from zipfile import BadZipFile

class xlsxFileProcessor(FileProcessorStrategy):
    def __init__(self, file_path: str) -> None:
        super().__init__(file_path)
        self.metadata = self._default_metadata()


    def _default_metadata(self) -> dict:
        return {
            "active_sheet": None,
            "sheet_names": None,
            "data": None,
            "last_modified_by": None,
            "creator": None,
            "has_password": False
        }

    
    def process(self) -> None:
        try:
            exceldoc = load_workbook(self.file_path)
            self.metadata.update({"active_sheet": exceldoc.active})
            self.metadata.update({"sheet_names": exceldoc.sheetnames})
            self.metadata.update({"data":self.read_all_data(exceldoc)})
            self.metadata.update({"last_modified_by": exceldoc.properties.lastModifiedBy})
            self.metadata.update({"creator": exceldoc.properties.creator})
        except BadZipFile:
            self.metadata['has_password'] = True
   
    def save(self, output_path: str = None) -> None:
        exceldoc = load_workbook(self.file_path)
        cp = exceldoc.properties
        # Update the core properties (metadata)
        cp.creator = self.metadata.get('creator', cp.creator)
        cp.last_modified_by = self.metadata.get('last_modified_by', cp.lastModifiedBy)
        
        save_path = output_path or self.file_path
        exceldoc.save(save_path)
 
    @staticmethod
    def read_all_data(exceldoc):
        data = {}
        for sheet_name in exceldoc.sheetnames:
            sheet = exceldoc[sheet_name]
            sheet_data = []
            for row in sheet.iter_rows(values_only=True):
                sheet_data.append(row)
            data[sheet_name] = sheet_data
        return data
